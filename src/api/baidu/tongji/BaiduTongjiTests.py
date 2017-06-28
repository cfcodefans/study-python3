import logging
import unittest
import traceback as tb

import json
import uuid
import math
import gzip
import rsa
import io
import requests
import openpyxl

from openpyxl.workbook import *
from openpyxl.worksheet import *
from openpyxl.cell import *
from openpyxl.styles import *
from openpyxl.styles.fills import *
from openpyxl.styles.colors import *
from typing import *

logging.basicConfig(format="%(asctime)s\t%(levelname)s\t%(message)s", level=logging.INFO)

UUID: str = "1q2w3e4r"
# USERNAME: str = "EUR-ACADEMIA"
# PASSWORD: str = "Academia2017"
# TOKEN: str = "8cea12f7dfe48e4f7d4f852cb7205ae6"

USERNAME: str = "EUR-MAGNETI"
PASSWORD: str = "Magneti2017"
TOKEN: str = "99a07f17f23b0ac7036bee7336bc858a"


def load_pub_key():
    with open("api_pub.key") as publicKeyFile:
        pk: str = publicKeyFile.read()
        # logging.info(pk)
        return rsa.PublicKey.load_pkcs1_openssl_pem(pk)


LOGIN_URL: str = "https://api.baidu.com/sem/common/HolmesLoginService"
API_URL: str = "https://api.baidu.com/json/tongji/v1/ReportService"
UUID: str = str(uuid.uuid1())
ACCOUNT_TYPE: str = "1"
PUB_KEY: rsa.PublicKey = load_pub_key()


def encrypt(data: bytes) -> str:
    n = int(math.ceil(len(data) * 1.0 / 117))
    ret: bytes = b''
    for i in range(n):
        gzdata: str = data[i * 117:(i + 1) * 117]
        ret += rsa.encrypt(gzdata, PUB_KEY)

    return ret


def gzencode(data: str) -> bytes:
    return gzip.compress(str.encode(data), 9)


def gzdecode(data: bytes) -> str:
    return gzip.decompress(data)


def pre_login():
    data = {'username': USERNAME,
            'token': TOKEN,
            'functionName': 'preLogin',
            'uuid': UUID,
            'request': {'osVersion': 'windows', 'deviceType': 'pc', 'clientVersion': '1.0'},
            }

    headers = get_headers()

    # 压缩
    post_data = gzencode(json.dumps(data))
    # 加密
    post_data = encrypt(post_data)

    resp: requests.Response = requests.post(LOGIN_URL, data=post_data, headers=headers)
    logging.info(resp.content[:8])
    logging.info(resp.content[8:])
    ret: any = json.loads(gzdecode(resp.content[8:]))
    logging.info('pre_login:\t\n' + str(ret))


def do_login() -> dict:
    data = {'username': USERNAME,
            'token': TOKEN,
            'functionName': 'doLogin',
            'uuid': UUID,
            'request': {'password': PASSWORD}
            }

    headers = get_headers()

    # 压缩
    post_data: bytes = gzencode(json.dumps(data))
    # 加密
    post_data: str = encrypt(post_data)
    # post
    resp: requests.Response = requests.post(LOGIN_URL, data=post_data, headers=headers)
    ret: any = json.loads(gzdecode(resp.content[8:]))
    if ret['retcode'] == 0:
        print(u'dologin:', ret['retmsg'])
        print(ret['ucid'])
        print(ret['st'])
    return ret


def get_headers() -> dict:
    headers = {'UUID': UUID, 'account_type': ACCOUNT_TYPE,
               'Content-Type': 'data/gzencode and rsa public encrypt;charset=UTF-8'
               }
    return headers


def get_site_list(ucid: str, st: str) -> list:
    url = API_URL + '/getSiteList'
    headers = {'UUID': UUID, 'USERID': ucid, 'Content-Type': 'data/json;charset=UTF-8'}
    data = {'header': {'username': USERNAME,
                       'password': st,
                       'token': TOKEN,
                       'account_type': ACCOUNT_TYPE, },
            'body': None, }
    post_data: str = json.dumps(data)
    resp: requests.Response = requests.post(url, data=post_data, headers=headers)
    print(resp.json())
    return resp.json()['body']['data'][0]['list']


def get_data(ucid: str, st: str, para: dict) -> dict:
    url = API_URL + '/getData'
    headers = {'UUID': UUID, 'USERID': ucid, 'Content-Type': 'data/json;charset=UTF-8'}
    data = {'header': {'username': USERNAME,
                       'password': st,
                       'token': TOKEN,
                       'account_type': ACCOUNT_TYPE, },
            'body': para, }

    post_data = json.dumps(data)
    resp: requests.Response = requests.post(url, data=post_data, headers=headers)
    print("resp.raw\n\t" + str(resp.json()))
    return resp.json()['body']


def write_table_to_excel(path: str,
                         title: str,
                         headers: list,
                         rows: List[list]):
    from openpyxl.utils import (
        get_column_letter,
        column_index_from_string,
    )

    wb: Workbook = Workbook()
    ws: Worksheet = wb.worksheets[0]
    ws.title = title.replace('/', '-')

    if len(rows) == 0:
        wb.save(path)
        return

    col_widths: list[int] = [len(header) for header in headers]
    for col in range(1, len(headers) + 1):
        cell: Cell = ws.cell(row=1, column=col, value=headers[col - 1])
        cell.fill = PatternFill(start_color='000000FF', end_color='000000FF', fill_type='solid')

    for row in range(2, len(rows) + 2):
        for col in range(1, len(headers)):
            value: str = str(rows[row - 2][col - 1])
            ws.cell(row=row, column=col, value=value)
            col_widths[col - 1] = max([col_widths[col - 1], len(value)])

    for col in range(1, len(col_widths) + 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1] + 2

    wb.save(path)

    pass


def write_result_to_excel(path: str, data: dict):
    result: dict = data["data"][0]["result"]
    timespan: str = result["timeSpan"][0]

    # ExcelWriter,里面封装好了对Excel的写操作
    from openpyxl.writer.excel import ExcelWriter
    # get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B
    from openpyxl.utils import (
        get_column_letter,
        column_index_from_string,
    )

    wb: Workbook = Workbook()
    # ew: ExcelWriter = ExcelWriter(wb)

    ws: Worksheet = wb.worksheets[0]
    ws.title = timespan.replace('/', '-')

    fields: list = result["fields"]
    # fields.insert(0, 'date')

    col_widths: list = [len(field) for field in fields]

    for col in range(1, len(fields) + 1):
        cell: Cell = ws.cell(row=1, column=col, value=fields[col - 1])
        cell.fill = PatternFill(start_color='000000FF', end_color='000000FF', fill_type='solid')

    items: list = result["items"]
    item_dates: list = [date_row for date_row in items[0]]
    item_values: list = [value_row for value_row in items[1]]
    item_date_values: list = [date_values[0] + date_values[1] for date_values in zip(item_dates, item_values)]
    row_width: int = len(item_date_values[0])
    for row in range(2, len(item_dates) + 1):
        for col in range(1, row_width):
            value: str = str(item_date_values[row - 2][col - 1])
            ws.cell(row=row, column=col, value=value)
            col_widths[col - 1] = max([col_widths[col - 1], len(value)])

    for col in range(1, len(col_widths) + 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1] + 2

    wb.save(path)

    pass


class BaiduTongjiTests(unittest.TestCase):
    # def test_pub_key(self):
    #     logging.info(str(PUB_KEY))

    # def test_encrypt(self):
    #     data: str = "1234567890" * 10 + "1234567"
    #     logging.info(data)
    #     logging.info(encrypt(data).hex())

    def test_result_excel(self):
        # raw = '{"data": [{"result": {"total": 32, "items": [[["2017/06/01"], ["2017/05/31"], ["2017/05/30"], ["2017/05/29"], ["2017/05/28"], ["2017/05/27"], ["2017/05/26"], ["2017/05/25"], ["2017/05/24"], ["2017/05/23"], ["2017/05/22"], ["2017/05/21"], ["2017/05/20"], ["2017/05/19"], ["2017/05/18"], ["2017/05/17"], ["2017/05/16"], ["2017/05/15"], ["2017/05/14"], ["2017/05/13"], ["2017/05/12"], ["2017/05/11"], ["2017/05/10"], ["2017/05/09"], ["2017/05/08"], ["2017/05/07"], ["2017/05/06"], ["2017/05/05"], ["2017/05/04"], ["2017/05/03"], ["2017/05/02"], ["2017/05/01"]], [[1, 1, 120], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], [100, 1, 342], ["--", "--", "--"], [328, 1, 1144], [20, 5, 941], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"], ["--", "--", "--"]], [], []], "timeSpan": ["2017/05/01 - 2017/06/01"], "sum": [[449, 8, 789], []], "offset": 0, "pageSum": [[449, 8, 789], [], []], "fields": ["simple_date_title", "pv_count", "visitor_count", "avg_visit_time"]}}]}'
        raw: dict = {'header': {'desc': 'success', 'failures': [], 'oprs': 1, 'succ': 1, 'oprtime': 0, 'quota': 1,
                                'rquota': 1990, 'status': 0}, 'body': {'data': [{'result': {'total': 6, 'items': [
            [['2017/06/01 - 2017/06/01'], ['2017/05/01 - 2017/05/31'], ['2017/04/01 - 2017/04/30'],
             ['2017/03/01 - 2017/03/31'], ['2017/02/01 - 2017/02/28'], ['2017/01/01 - 2017/01/31']],
            [[86], [1868], [1589], [1462], [919], [687]], [], []], 'timeSpan': ['2017/01/01 - 2017/06/01'],
                                                                                            'sum': [[6611], []],
                                                                                            'offset': 0,
                                                                                            'pageSum': [[6611], [], []],
                                                                                            'fields': [
                                                                                                'simple_date_title',
                                                                                                'pv_count']}}]}}
        mock_data: dict = raw['body']
        result = mock_data["data"][0]["result"]
        timespan: str = result["timeSpan"][0]
        logging.info("timespan:\t" + timespan)
        fields: list = result["fields"]
        logging.info("\t".join(fields))
        headers: list = [dates[0] for dates in result['items'][0]]
        logging.info("\t".join(headers))
        rows: list = [str(cols[0]) for cols in result['items'][1]]
        logging.info("\t".join(rows))


        # write_result_to_excel("test.xlsx", mock_data)

    def test_month_visits(self):
        pre_login()
        login_info: dict = do_login()

        ucid: str = str(login_info["ucid"])
        st: str = str(login_info["st"])

        site_list: list = get_site_list(ucid, st)
        logging.info("site list:\n" + "\n".join([str(site) for site in site_list]))

        site_id: str = site_list[0]["site_id"]
        para: dict = {'site_id': site_id,  # 站点ID
                      'method': 'trend/time/a',  # 趋势分析报告
                      'start_date': '20170101',  # 所查询数据的起始日期
                      'end_date': '20170601',  # 所查询数据的结束日期
                      'metrics': 'pv_count',  # 所查询指标为PV和UV
                      'max_results': '0',  # 返回所有条数
                      'gran': 'month',  # 粒度
                      'source': 'through'
                      }

        logging.info("\n")
        data_may: dict = get_data(ucid, st, para)
        _result: dict = data_may['data'][0]['result']
        timespan: str = _result["timeSpan"][0]

        search_visit_items: list = _result['items']
        headers: list = [_result['fields'][0]] + [dates[0] for dates in search_visit_items[0]]

        direct_visits: list = ['direct visits'] + [str(cols[0]) for cols in search_visit_items[1]]

        para['source'] = "search,0"
        data_may = get_data(ucid, st, para)
        _result = data_may['data'][0]['result']
        search_visit_items = _result['items']
        search_visits: list = ['visits from search engine'] + [str(cols[0]) for cols in search_visit_items[1]]

        para['source'] = "link"
        data_may = get_data(ucid, st, para)
        _result = data_may['data'][0]['result']
        search_visit_items = _result['items']
        links_visits: list = ['visits from links'] + [str(cols[0]) for cols in search_visit_items[1]]

        write_table_to_excel("month-visits.xlsx", timespan, headers, [direct_visits, search_visits, links_visits])


if __name__ == "__main__":
    unittest.main()
